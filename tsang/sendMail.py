import os
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import getpass
import datetime

def get_user_credentials():
    sender_email = input("Enter your Outlook email address: ")
    sender_password = getpass.getpass("Enter your app password (generated from Outlook): ")
    return sender_email, sender_password

def get_script_directory():
    return os.path.dirname(os.path.realpath(__file__))

# Manual input of sender's email and app password
sender_email, sender_password = get_user_credentials()

# 获取脚本所在目录
script_directory = get_script_directory()

# 构造Excel文件路径
excel_file_path = os.path.join(script_directory, 'SendMail.xlsx')

# 邮件内容
msg = MIMEMultipart()
msg['Subject'] = '邮件主题'
msg['From'] = sender_email

# 电子邮件正文
body = '邮件正文'
msg.attach(MIMEText(body, 'plain'))

# 发送邮件
smtp_server = 'smtp-mail.outlook.com'  # Outlook SMTP server
smtp_port = 587

smtpObj = smtplib.SMTP(smtp_server, smtp_port)

try:
    smtpObj.starttls()
    smtpObj.ehlo()
    smtpObj.login(sender_email, sender_password)

    num_of_tasks = 0

    if os.path.exists(excel_file_path):
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active
        num_of_tasks = sheet.max_row - 1

        print(f'当前时间：{datetime.datetime.now()}')
        print(f'共找到{num_of_tasks}条发送任务。')

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            subject = row[0]
            body = row[1]
            attachment_path = row[2]
            to_recipients = row[3].split(',')
            cc_recipients = row[4].split(',') if row[4] else []
            bcc_recipients = row[5].split(',') if row[5] else []

            # 邮件内容
            msg = MIMEMultipart()
            msg['Subject'] = subject
            msg['From'] = sender_email
            msg['To'] = ', '.join(to_recipients)
            msg['Cc'] = ', '.join(cc_recipients)
            msg['Bcc'] = ', '.join(bcc_recipients)

            # 电子邮件正文
            msg.attach(MIMEText(body, 'plain'))

            # 附件
            if attachment_path:
                attachment_path = os.path.join(script_directory, attachment_path)
                filename = os.path.basename(attachment_path)
                attachment = open(attachment_path, 'rb').read()
                part = MIMEApplication(attachment, Name=filename)
                part['Content-Disposition'] = f'attachment; filename="{filename}"'
                msg.attach(part)

            smtpObj.sendmail(sender_email, to_recipients + cc_recipients + bcc_recipients, msg.as_string())
            print(f'第{i}条已成功发送，收件人{to_recipients}，抄送人{cc_recipients}，密送人{bcc_recipients}')

except smtplib.SMTPAuthenticationError:
    print("Error: Outlook邮箱地址或应用程序专用密码有误，请检查你的账户信息。")

except Exception as e:
    print(f"Error: {e}")

finally:
    smtpObj.quit()
